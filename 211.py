import csv
import re
from itertools import islice
from typing import List, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

currency_to_rub = {'AZN': 35.68, 'BYR': 23.91, 'EUR': 59.9, 'GEL': 21.74, 'KGS': 0.76, 'KZT': 0.13, 'RUR': 1, 'UAH': 1.64, 'USD': 60.66, 'UZS': 0.0055}

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

class DataSet:
    def __init__(self, file_name, vacancies_objects, **kwargs):
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

class Vacancy:
    def __init__(self, name, salary, area_name, published_at, **kwargs):
        self.name: str = name
        self.area_name: str = area_name
        self.salary: Salary = salary
        self.published_at: datetime = datetime.strptime(published_at.replace('T', ' '), '%Y-%m-%d %H:%M:%S+%f')

class Salary:
    def __init__(self, salary_from, salary_to, salary_currency, **kwargs):
        self.salary_from: SalaryFloatItem = SalaryFloatItem(salary_from)
        self.salary_to: SalaryFloatItem = SalaryFloatItem(salary_to)
        self.salary_currency: str = salary_currency

        if salary_currency != 'RUR':
            self.salary_from = currency_to_rub[salary_currency] * self.salary_from
            self.salary_to = currency_to_rub[salary_currency] * self.salary_to

    def __repr__(self):
        return f'{self.salary_from} - {self.salary_to} ({self.salary_currency}) {self.salary_gross}'

    def __len__(self):
        return len(str(self))

class SalaryFloatItem(float):
    def __init__(self, value):
        self = value

    def __repr__(self):
        return f'{int(self):,}'.replace(',', ' ')


class Report:
    def __init__(self, prof_name, salaries, vacancies, salaries_prof, vacancies_prof):
        self.prof_name = prof_name
        self.salaries = salaries
        self.vacancies = vacancies
        self.salaries_prof = salaries_prof
        self.vacancies_prof = vacancies_prof

    def generate_excel(self):
        wb = Workbook()
        ws = wb.active
        row_names = ('Год', 'Средняя зарплата', f'Средняя зарплата - {self.prof_name}', 'Количество вакансий', f'Количество вакансий - {self.prof_name}')
        ws.append(row_names)

        max_len_dict = max((self.salaries, self.vacancies, self.salaries_prof, self.vacancies_prof), key=lambda i: len(i))

        # задаем колонкам длину и жирный шрифт для заголовков
        for i, cell in enumerate(ws['A1':f'{get_column_letter(len(row_names))}1'][0], 1):
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(i)].width = len(cell.value) + 2

        # устанавливаем рамки для таблицы
        for col in ws.iter_cols(max_col=len(row_names), max_row=len(max_len_dict)+1):
            for cell in col:
                cell.border = thin_border

        # в первый столбец выводит года из самого длинного словаря (чтобы для всех словарей хватило строк с годами)
        for row, year in enumerate(max_len_dict.keys(), start=2):
            ws.cell(row=row, column=1, value=year)

        # заполняем таблицу
        for column, dict_ in enumerate((self.salaries, self.salaries_prof, self.vacancies, self.vacancies_prof), start=2):
            for row, value in enumerate(dict_.values(), start=2):
                ws.cell(row=row, column=column, value=value)

        wb.save("report.xlsx")

def csv_reader(file_name: str) -> Tuple[List[str], List[str]]:
    with open(file_name, 'r', encoding='utf-8', newline='') as file:
        titles = re.sub('\n|\r|\ufeff', '', file.readline()).split(',')
        data = [elem for elem in list(csv.reader(file)) if all(map(lambda x: len(x) > 0, elem)) and len(elem) == len(titles)]
       
        file.seek(0)
        if file.read() == '':
            return 'Пустой файл'
        if len(data) == 0:
            return 'Нет данных'

    return titles, data

def format_value(dict_object: dict, key: str, value: str) -> None:
    value = re.sub('\r', '', value)
    value = re.sub(r'<[^>]+>', '', value, flags=re.S)
    value = '\n'.join(map(lambda i: i.strip(), value.split('\n'))) if '\n' in value else ' '.join(value.strip().split())    
    dict_object[key] = value

def csv_filer(titles: List[str], data: List[str]) -> List[Vacancy]:
    vacancies_objects = []

    for vacancy_data in data:
        vacancy = {key: ... for key in ('name', 'area_name', 'published_at')}
        salary = {key: ... for key in ('salary_from', 'salary_to', 'salary_currency')}
        for key, value in zip(titles, vacancy_data):
            format_value(salary if 'salary' in key else vacancy, key, value)

        vacancy['salary'] = Salary(**salary)
        vacancies_objects.append(Vacancy(**vacancy))

    return vacancies_objects

def add_data(dict_object: dict, key: str, average_salary: float, add_empty: bool) -> None:
    if add_empty:
        dict_object[key] = {'salary': [], 'count': 0}
    dict_object[key]['salary'].append(average_salary)
    dict_object[key]['count'] += 1

def calculate_average_salary(dict_object: dict) -> None:
    for key in dict_object:
        dict_object[key]['salary'] = int(sum(dict_object[key]['salary']) / len(dict_object[key]['salary']))

def print_statistics(vacancies_data: List[Vacancy], prof_name: str) -> None:
    total_data = {}
    prof_data = {}
    cities = {}

    for vacancy in vacancies_data:
        average_salary = (vacancy.salary.salary_from + vacancy.salary.salary_to) / 2
        # статистика городов
        add_data(cities, vacancy.area_name, average_salary, vacancy.area_name not in cities.keys())
        # зарплаты и вакансии
        add_data(total_data, vacancy.published_at.year, average_salary, vacancy.published_at.year not in total_data.keys())
        # зарплаты и вакансии для профессии
        if prof_name in vacancy.name:
            add_data(prof_data, vacancy.published_at.year, average_salary, vacancy.published_at.year not in prof_data.keys())

    for dict_ in (total_data, prof_data, cities):
        calculate_average_salary(dict_)

    # убираем все города, в которых количество вакансий меньше 1% от общего числа вакансий
    cities = {k: v for k, v in cities.items() if (lambda v: 1 if v >= 0.75 else 0)(cities[k]['count'] / len(vacancies_data) * 100) >= 1}

    salaries = {year: total_data[year]["salary"] for year in total_data}
    vacancies = {year: total_data[year]["count"] for year in total_data}
    salaries_prof = {year: prof_data[year]["salary"] for year in prof_data}
    vacancies_prof = {year: prof_data[year]["count"] for year in prof_data}

    print('Динамика уровня зарплат по годам:', salaries)
    print('Динамика количества вакансий по годам:', vacancies)
    print('Динамика уровня зарплат по годам для выбранной профессии:', salaries_prof)
    print('Динамика количества вакансий по годам для выбранной профессии:', vacancies_prof)
    top10 = dict(islice({k: v for k, v in sorted(cities.items(), key=lambda item: item[1]['salary'], reverse=True)}.items(), 10))
    print('Уровень зарплат по городам (в порядке убывания):', {k: top10[k]['salary'] for k in top10})
    top10 = dict(islice({k: v for k, v in sorted(cities.items(), key=lambda item: item[1]['count'], reverse=True)}.items(), 10))
    print('Доля вакансий по городам (в порядке убывания):', {k: float(f"{(top10[k]['count'] / len(vacancies_data)):.4f}") for k in top10})

    Report(prof_name, salaries, vacancies, salaries_prof, vacancies_prof).generate_excel()

def get_input():
    file_name = input('Введите название файла: ')
    prof_name = input('Введите название профессии: ')

    csv_data = csv_reader(file_name)
    if isinstance(csv_data, str):
        return print(csv_data)

    print_statistics(csv_filer(*csv_data), prof_name)


get_input()